"""
PDF conversion services.
"""
import time
from pathlib import Path
from pdf2docx import Converter as PDF2DOCXConverter
from docx2pdf import convert as docx2pdf_convert
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from pptx import Presentation

from apps.tools.utils.base_converter import BaseConverter, ConversionError
from apps.tools.utils.converter_factory import register_converter


@register_converter('pdf_to_docx')
class PDFToDocxConverter(BaseConverter):
    """
    Converter for PDF to DOCX format using pdf2docx library.
    """
    
    ALLOWED_INPUT_TYPES = ['application/pdf']
    ALLOWED_INPUT_EXTENSIONS = ['pdf']
    MAX_FILE_SIZE_MB = 50
    OUTPUT_EXTENSION = 'docx'
    
    def convert(self, input_path, output_path):
        """
        Convert PDF file to DOCX format.
        
        Args:
            input_path: Path to input PDF file
            output_path: Path where output DOCX file should be saved
            
        Returns:
            dict: Conversion result with status and metadata
        """
        start_time = time.time()
        self.log_conversion_start(input_path, output_path)
        
        try:
            # Validate input file
            self.validate_file(input_path)
            
            # Perform conversion
            cv = PDF2DOCXConverter(input_path)
            cv.convert(output_path)
            cv.close()
            
            duration = time.time() - start_time
            self.log_conversion_success(input_path, output_path, duration)
            
            return {
                'status': 'success',
                'output_path': output_path,
                'duration': duration,
                'input_info': self.get_file_info(input_path),
                'output_info': self.get_file_info(output_path),
            }
            
        except Exception as e:
            self.log_conversion_error(input_path, e)
            raise ConversionError(f"PDF to DOCX conversion failed: {str(e)}")


@register_converter('docx_to_pdf')
class DocxToPDFConverter(BaseConverter):
    """
    Converter for DOCX to PDF format using LibreOffice (Linux/Mac) or Word COM (Windows).
    """
    
    ALLOWED_INPUT_TYPES = [
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/msword'  # Support .doc files
    ]
    ALLOWED_INPUT_EXTENSIONS = ['docx', 'doc']
    MAX_FILE_SIZE_MB = 50
    OUTPUT_EXTENSION = 'pdf'
    
    def _convert_with_word(self, input_path, output_path):
        """
        Convert DOCX/DOC to PDF using Microsoft Word COM automation (Windows only).
        
        Args:
            input_path: Path to input DOCX/DOC file
            output_path: Path where output PDF file should be saved
            
        Raises:
            ConversionError: If Word conversion fails
        """
        import os
        
        try:
            import win32com.client
        except ImportError:
            raise ConversionError(
                "win32com is not available. This method only works on Windows with pywin32 installed."
            )
        
        word = None
        doc = None
        
        try:
            # Convert paths to absolute paths
            abs_input = os.path.abspath(input_path)
            abs_output = os.path.abspath(output_path)
            
            # Initialize Word application
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            
            # Open document
            doc = word.Documents.Open(abs_input)
            
            # Save as PDF (format 17 is PDF)
            doc.SaveAs(abs_output, FileFormat=17)
            
            self.logger.info(f"Word COM conversion successful: {input_path} -> {output_path}")
            
        except Exception as e:
            error_msg = f"Word COM automation failed: {str(e)}"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
            
        finally:
            # Clean up COM objects
            try:
                if doc:
                    doc.Close()
            except Exception as e:
                self.logger.warning(f"Error closing document: {str(e)}")
            
            try:
                if word:
                    word.Quit()
            except Exception as e:
                self.logger.warning(f"Error quitting Word: {str(e)}")
    
    def _convert_with_libreoffice(self, input_path, output_path):
        """
        Convert DOCX/DOC to PDF using LibreOffice command-line (cross-platform).
        
        Args:
            input_path: Path to input DOCX/DOC file
            output_path: Path where output PDF file should be saved
            
        Raises:
            ConversionError: If LibreOffice conversion fails
        """
        import os
        import subprocess
        from apps.tools.utils.platform_utils import get_libreoffice_path
        
        # Get LibreOffice executable path
        libreoffice_path = get_libreoffice_path()
        if not libreoffice_path:
            raise ConversionError(
                "LibreOffice is not installed or not found. "
                "Please install LibreOffice to enable DOCX/DOC to PDF conversion.\n"
                "Installation instructions:\n"
                "  - Ubuntu/Debian: sudo apt-get install libreoffice\n"
                "  - RHEL/CentOS: sudo yum install libreoffice\n"
                "  - Windows: Download from https://www.libreoffice.org/download/\n"
                "  - Mac: Download from https://www.libreoffice.org/download/"
            )
        
        try:
            # Convert to absolute paths
            abs_input = os.path.abspath(input_path)
            abs_output = os.path.abspath(output_path)
            output_dir = os.path.dirname(abs_output)
            
            # Calculate timeout based on file size
            file_size_mb = os.path.getsize(abs_input) / (1024 * 1024)
            timeout = max(300, int(300 + (file_size_mb * 3)))
            
            # Build LibreOffice command
            command = [
                libreoffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                abs_input
            ]
            
            self.logger.info(f"Running LibreOffice conversion: {' '.join(command)}")
            
            # Run LibreOffice conversion
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False
            )
            
            # Check for errors
            if result.returncode != 0:
                error_msg = f"LibreOffice conversion failed with exit code {result.returncode}"
                if result.stderr:
                    error_msg += f"\nStderr: {result.stderr}"
                if result.stdout:
                    error_msg += f"\nStdout: {result.stdout}"
                
                self.logger.error(error_msg)
                raise ConversionError(error_msg)
            
            # LibreOffice creates the output file with the same name as input but .pdf extension
            input_filename = os.path.splitext(os.path.basename(abs_input))[0]
            libreoffice_output = os.path.join(output_dir, f"{input_filename}.pdf")
            
            # If LibreOffice output is different from expected output, rename it
            if libreoffice_output != abs_output and os.path.exists(libreoffice_output):
                if os.path.exists(abs_output):
                    os.remove(abs_output)
                os.rename(libreoffice_output, abs_output)
            
            # Verify output file exists
            if not os.path.exists(abs_output):
                raise ConversionError(
                    f"LibreOffice conversion completed but output file not found: {abs_output}"
                )
            
            self.logger.info(f"LibreOffice conversion successful: {input_path} -> {output_path}")
            
        except subprocess.TimeoutExpired:
            error_msg = f"LibreOffice conversion timed out after {timeout} seconds"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
        except Exception as e:
            if isinstance(e, ConversionError):
                raise
            error_msg = f"LibreOffice conversion failed: {str(e)}"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
    
    def convert(self, input_path, output_path):
        """
        Convert DOCX/DOC file to PDF format.
        
        Args:
            input_path: Path to input DOCX/DOC file
            output_path: Path where output PDF file should be saved
            
        Returns:
            dict: Conversion result with status and metadata
        """
        from apps.tools.utils.platform_utils import get_platform, is_powerpoint_available
        
        start_time = time.time()
        self.log_conversion_start(input_path, output_path)
        
        try:
            # Validate input file
            self.validate_file(input_path)
            
            # Determine conversion method based on platform
            current_platform = get_platform()
            
            if current_platform == 'windows':
                # On Windows, try Word COM first, then fall back to LibreOffice
                try:
                    self._convert_with_word(input_path, output_path)
                except ConversionError as e:
                    self.logger.warning(f"Word COM failed, trying LibreOffice: {str(e)}")
                    self._convert_with_libreoffice(input_path, output_path)
            else:
                # On Linux/Mac, use LibreOffice
                self._convert_with_libreoffice(input_path, output_path)
            
            duration = time.time() - start_time
            self.log_conversion_success(input_path, output_path, duration)
            
            return {
                'status': 'success',
                'output_path': output_path,
                'duration': duration,
                'input_info': self.get_file_info(input_path),
                'output_info': self.get_file_info(output_path),
            }
            
        except Exception as e:
            self.log_conversion_error(input_path, e)
            raise ConversionError(f"DOCX to PDF conversion failed: {str(e)}")


@register_converter('xlsx_to_pdf')
class XLSXToPDFConverter(BaseConverter):
    """
    Converter for XLSX to PDF format using openpyxl and ReportLab.
    """
    
    ALLOWED_INPUT_TYPES = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]
    ALLOWED_INPUT_EXTENSIONS = ['xlsx']
    MAX_FILE_SIZE_MB = 50
    OUTPUT_EXTENSION = 'pdf'
    
    def convert(self, input_path, output_path):
        """
        Convert XLSX file to PDF format.
        
        Args:
            input_path: Path to input XLSX file
            output_path: Path where output PDF file should be saved
            
        Returns:
            dict: Conversion result with status and metadata
        """
        start_time = time.time()
        self.log_conversion_start(input_path, output_path)
        
        try:
            # Validate input file
            self.validate_file(input_path)
            
            # Load workbook
            wb = load_workbook(input_path, data_only=True)
            
            # Create PDF
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Add sheet title
                elements.append(Paragraph(f"<b>{sheet_name}</b>", styles['Heading1']))
                elements.append(Spacer(1, 12))
                
                # Convert sheet data to table
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append([str(cell) if cell is not None else '' for cell in row])
                
                if data:
                    # Create table
                    table = Table(data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(elements)
            
            duration = time.time() - start_time
            self.log_conversion_success(input_path, output_path, duration)
            
            return {
                'status': 'success',
                'output_path': output_path,
                'duration': duration,
                'sheets_processed': len(wb.sheetnames),
                'input_info': self.get_file_info(input_path),
                'output_info': self.get_file_info(output_path),
            }
            
        except Exception as e:
            self.log_conversion_error(input_path, e)
            raise ConversionError(f"XLSX to PDF conversion failed: {str(e)}")


@register_converter('pptx_to_pdf')
class PPTXToPDFConverter(BaseConverter):
    """
    Converter for PPTX to PDF format using PowerPoint COM automation or LibreOffice.
    """
    
    ALLOWED_INPUT_TYPES = [
        'application/vnd.openxmlformats-officedocument.presentationml.presentation'
    ]
    ALLOWED_INPUT_EXTENSIONS = ['pptx']
    MAX_FILE_SIZE_MB = 100
    OUTPUT_EXTENSION = 'pdf'
    
    def _convert_with_powerpoint(self, input_path, output_path):
        """
        Convert PPTX to PDF using PowerPoint COM automation (Windows only).
        
        Args:
            input_path: Path to input PPTX file
            output_path: Path where output PDF file should be saved
            
        Raises:
            ConversionError: If PowerPoint conversion fails
        """
        import os
        
        try:
            import win32com.client
        except ImportError:
            raise ConversionError(
                "win32com is not available. This method only works on Windows with pywin32 installed."
            )
        
        powerpoint = None
        presentation = None
        
        try:
            # Convert to absolute paths (COM requires absolute paths)
            abs_input = os.path.abspath(input_path)
            abs_output = os.path.abspath(output_path)
            
            # Initialize PowerPoint application
            powerpoint = win32com.client.Dispatch("PowerPoint.Application")
            powerpoint.Visible = 0  # Run headless (0 = hidden, 1 = visible)
            
            # Open presentation without window
            presentation = powerpoint.Presentations.Open(
                abs_input,
                ReadOnly=True,
                Untitled=True,
                WithWindow=False
            )
            
            # Save as PDF (FileFormat=32 is PDF)
            presentation.SaveAs(abs_output, FileFormat=32)
            
            self.logger.info(f"PowerPoint COM conversion successful: {input_path} -> {output_path}")
            
        except Exception as e:
            error_msg = f"PowerPoint COM automation failed: {str(e)}"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
            
        finally:
            # Clean up COM objects
            try:
                if presentation:
                    presentation.Close()
            except Exception as e:
                self.logger.warning(f"Error closing presentation: {str(e)}")
            
            try:
                if powerpoint:
                    powerpoint.Quit()
            except Exception as e:
                self.logger.warning(f"Error quitting PowerPoint: {str(e)}")
    
    def _convert_with_libreoffice(self, input_path, output_path):
        """
        Convert PPTX to PDF using LibreOffice command-line (cross-platform).
        
        Args:
            input_path: Path to input PPTX file
            output_path: Path where output PDF file should be saved
            
        Raises:
            ConversionError: If LibreOffice conversion fails
        """
        import os
        import subprocess
        from apps.tools.utils.platform_utils import get_libreoffice_path
        
        # Get LibreOffice executable path
        libreoffice_path = get_libreoffice_path()
        if not libreoffice_path:
            raise ConversionError(
                "LibreOffice is not installed or not found. "
                "Please install LibreOffice to enable PPTX to PDF conversion.\n"
                "Installation instructions:\n"
                "  - Ubuntu/Debian: sudo apt-get install libreoffice\n"
                "  - RHEL/CentOS: sudo yum install libreoffice\n"
                "  - Windows: Download from https://www.libreoffice.org/download/\n"
                "  - Mac: Download from https://www.libreoffice.org/download/"
            )
        
        try:
            # Convert to absolute paths
            abs_input = os.path.abspath(input_path)
            abs_output = os.path.abspath(output_path)
            output_dir = os.path.dirname(abs_output)
            
            # Calculate timeout based on file size (300 seconds base, scale with file size)
            file_size_mb = os.path.getsize(abs_input) / (1024 * 1024)
            timeout = max(300, int(300 + (file_size_mb * 3)))  # Add 3 seconds per MB
            
            # Build LibreOffice command
            command = [
                libreoffice_path,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                abs_input
            ]
            
            self.logger.info(f"Running LibreOffice conversion: {' '.join(command)}")
            
            # Run LibreOffice conversion
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=timeout,
                check=False  # Don't raise exception on non-zero exit code
            )
            
            # Check for errors
            if result.returncode != 0:
                error_msg = f"LibreOffice conversion failed with exit code {result.returncode}"
                if result.stderr:
                    error_msg += f"\nStderr: {result.stderr}"
                if result.stdout:
                    error_msg += f"\nStdout: {result.stdout}"
                
                self.logger.error(error_msg)
                raise ConversionError(error_msg)
            
            # LibreOffice creates the output file with the same name as input but .pdf extension
            # We need to rename it to the expected output path
            input_filename = os.path.splitext(os.path.basename(abs_input))[0]
            libreoffice_output = os.path.join(output_dir, f"{input_filename}.pdf")
            
            # If LibreOffice output is different from expected output, rename it
            if libreoffice_output != abs_output and os.path.exists(libreoffice_output):
                if os.path.exists(abs_output):
                    os.remove(abs_output)
                os.rename(libreoffice_output, abs_output)
            
            # Verify output file exists
            if not os.path.exists(abs_output):
                raise ConversionError(
                    f"LibreOffice conversion completed but output file not found: {abs_output}"
                )
            
            self.logger.info(f"LibreOffice conversion successful: {input_path} -> {output_path}")
            
            # Log stdout/stderr for debugging if present
            if result.stdout:
                self.logger.debug(f"LibreOffice stdout: {result.stdout}")
            if result.stderr:
                self.logger.debug(f"LibreOffice stderr: {result.stderr}")
                
        except subprocess.TimeoutExpired as e:
            error_msg = f"LibreOffice conversion timed out after {timeout} seconds"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
        except Exception as e:
            if isinstance(e, ConversionError):
                raise
            error_msg = f"LibreOffice conversion failed: {str(e)}"
            self.logger.error(error_msg)
            raise ConversionError(error_msg)
    
    def convert(self, input_path, output_path):
        """
        Convert PPTX file to PDF format.
        
        Args:
            input_path: Path to input PPTX file
            output_path: Path where output PDF file should be saved
            
        Returns:
            dict: Conversion result with status and metadata
        """
        from apps.tools.utils.platform_utils import get_platform, is_powerpoint_available
        
        start_time = time.time()
        self.log_conversion_start(input_path, output_path)
        
        try:
            # Validate input file
            self.validate_file(input_path)
            
            # Determine conversion method based on platform
            current_platform = get_platform()
            
            if current_platform == 'windows' and is_powerpoint_available():
                # On Windows with PowerPoint, use COM automation
                try:
                    self._convert_with_powerpoint(input_path, output_path)
                except ConversionError as e:
                    self.logger.warning(f"PowerPoint COM failed, trying LibreOffice: {str(e)}")
                    self._convert_with_libreoffice(input_path, output_path)
            else:
                # On Linux/Mac or Windows without PowerPoint, use LibreOffice
                self._convert_with_libreoffice(input_path, output_path)
            
            duration = time.time() - start_time
            self.log_conversion_success(input_path, output_path, duration)
            
            # Get slide count for metadata
            try:
                prs = Presentation(input_path)
                slides_count = len(prs.slides)
            except:
                slides_count = None
            
            return {
                'status': 'success',
                'output_path': output_path,
                'duration': duration,
                'slides_processed': slides_count,
                'input_info': self.get_file_info(input_path),
                'output_info': self.get_file_info(output_path),
            }
            
        except Exception as e:
            self.log_conversion_error(input_path, e)
            raise ConversionError(f"PPTX to PDF conversion failed: {str(e)}")
